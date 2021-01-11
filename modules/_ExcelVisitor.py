from os import path, sendfile
from openpyxl import load_workbook


class ExcelVisitor:
    PATH = "src/excel/"

    def __init__(self,path=None, sheet_name=None):
        """
        You could either open the sheet here or in the open method
        """
        if(path is not None):
            self.open(path, sheet_name)

    def open(self, path, sheet_name=None):
        self.workbook = load_workbook(path, data_only=True)
        if(len(self.workbook.sheetnames)==0): raise("Excel sheet is empty")
        if(sheet_name is None): sheet_name=(self.workbook.sheetnames)[0]
        self.worksheet = self.workbook[sheet_name]
    
    def close(self, path):
        self.workbook.save(path)

    def read_row(self, row):
        """
        row:: int
        example:  1, 2, ,3 
        """
        if(self.worksheet is None): raise("Worksheet not initialized !")
        table_row = self.worksheet[row]
        return [cell.value for cell in table_row]
    
    def read_col(self, col):
        """
        col
        example: A, B, C ...
        """
        if(self.worksheet is None): raise("Worksheet not initialized !")
        table_col = self.worksheet[str(col)]
        return [cell.value for cell in table_col]

    def read_all(self):
        """
        Return all the value in table as 2d array (might have some empty entries
        which will be represented as none in the array

        Exmaple: 
            table : | 1 |   | 2 |
                    | 1 | 2 | 2 |
                    | 1 | 2 | 2 |
            return: [[1, None, 2], [1,2,2], [1,2,2]]

        """

        if(self.worksheet is None): raise("Worksheet not initialized !")

        temp_table = [] 

        # Traverse all the rows in a single sheet/table
        for row in self.worksheet:
            temp_rowsList = []
            has_non_empty_element = False

            # Traverse all the cells in a single row
            for cell in row: 

                # Use placeholder None in case empty
                if(cell.value is not None):
                    has_non_empty_element = True
                    temp_rowsList.append(str(cell.value))
                else:
                    temp_rowsList.append(None)

            temp_table.append(temp_rowsList)

        return temp_table

    def get_titles(self):
        """
        Get first row in the table while also eliminating 
        all the none entries in the row (will only have a set of actual titles)
        """
        if(self.worksheet is None): raise("Worksheet not initialized !")
        title_row = self.read_row(1)
        title_row_last = title_row[-1]
        while(title_row_last is None):
            title_row = title_row[:-1]
            title_row_last = title_row[-1]
        return title_row

    def get_infoTable(self, keep_title=False):
        info_cols  = self.get_titles()
        raw_table  = self.read_all()
        info_table = []
        
        # Remove operator columns 
        for old_row in raw_table:
            new_row = old_row[:len(info_cols)]
            info_table.append(new_row)

        # Remove any row that is all none 
        # i.e.[None, None, None, ..., None]
        last_row   = info_table[-1]
        last_row_exist_value = False
        for cell in last_row:
            if(cell is not None): last_row_exist_value = True
        while(not last_row_exist_value):
            info_table = info_table[:-1]
            last_row   = info_table[-1]
            for cell in last_row:
                if(cell is not None): last_row_exist_value = True

        # Remove the title row?
        if(keep_title): return info_table
        else: return info_table[1:]
        
    def get_operTable(self):
        num_info_rows = len(self.get_infoTable())
        info_cols  = self.get_titles()
        raw_table  = self.read_all()
        oper_table = []
        
        try: 
            # Retain operator columns 
            for old_row in raw_table:
                oper_start_col = len(info_cols)
                new_row = old_row[oper_start_col: oper_start_col+2]
                oper_table.append(new_row)
        except:
            oper_table=[]

        # in case none of the oeprator are filled 
        if(len(oper_table) == 0):
            oper_table = [[None, None] for i in range(num_info_rows)]

        # Remove first row (which is where the title row is)
        # and set threshold depth to be the same as the info table
        return oper_table[1:num_info_rows+1] 



        
    











